import openpyxl


class excelUtilit:
    @staticmethod
    def writeDataInExpectedCell(excelFile,columnName, searchTerm,new_value):
        workbook = openpyxl.load_workbook(excelFile)
        sheet = workbook.active
        columnNumber=1
        for i in range(1,sheet.max_column+1):
            if sheet.cell(row=1,column=i)==columnName:
                columnNumber=i
                break

        for i in range(1, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                if sheet.cell(row=i, column=j).value == searchTerm:
                    sheet.cell(row=i, column=columnNumber).value = new_value
       #This To save The File
        workbook.save(excelFile)